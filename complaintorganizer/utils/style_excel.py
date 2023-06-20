from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import pandas as pd

def get_color(route):
    if any(substring in route for substring in ["HHRESZ", "HHRESK"]):
        return "FAD201"
    elif any(substring in route for substring in ["HHGFZC", "HHGFZZ"]):
        return "B4C424"
    elif "HHRESC" in route:
        return "C4B454"
    elif "HHPAP" in route:
        return "FBCEB1"
    elif "HHPMD" in route:
        return "87CEEB"
    elif "HHGLS" in route:
        return "FFAA33"
    else:
        return "FFFFFF"  # Default color

def create_styled_excel(df):
    color = get_color(df['route'].iloc[0])
    if df is not None:
        # Create a workbook and add a worksheet to it
        wb = Workbook()
        ws = wb.active

        # Define styles
        header_font = Font(name='Calibri', bold=True, color=Color('000000'), size=16)  # Make text color black
        header_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        subheader_font = Font(name='Calibri', bold=True, color=Color('000000'),
                              size=14)  # Larger font for subheaders
        column_name_font = Font(name='Calibri', bold=True, color=Color('000000'))  # Bold font for column names
        cell_font = Font(name='Calibri', color=Color('000000'))
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))  # Set all borders

        # Create the main header
        route_value = df['route'].iloc[0]
        ws.merge_cells('A1:F2')
        ws['A1'] = 'COMPLAINTS FOR ROUTE {}'.format(route_value)
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].border = thin_border

        # Remove the 'route_id' and 'complaint_date' columns
        df = df.drop(columns=['route'])

        # Reorder the columns
        df = df[['street_address', 'address_number', 'zip_code', 'city', 'modality', 'recipient', 'complaint_date']]

        # Convert 'complaint_date' to datetime type (date only)
        df['complaint_date'] = pd.to_datetime(df['complaint_date']).dt.date

        # Write column names to 3rd row
        for i, column_name in enumerate(df.columns, start=1):
            ws.cell(row=3, column=i, value=column_name)
            ws.cell(row=3, column=i).font = column_name_font
            ws.cell(row=3, column=i).border = thin_border

        # Write data to worksheet
        current_row = 4
        # Write data to worksheet
        df['week'] = pd.to_datetime(df['complaint_date']).dt.to_period('W')  # Add 'week' column for weekly grouping
        for _, week_group in df.groupby('week'):
            week_dates = sorted(list(week_group['complaint_date'].unique()))
            week_dates_str = ' + '.join(map(str, week_dates))  # Concatenate dates of the same week
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            cell = ws.cell(row=current_row, column=1, value=week_dates_str)
            cell.font = subheader_font
            cell.fill = header_fill
            cell.alignment = cell_alignment
            cell.border = thin_border
            current_row += 1

            week_group = week_group.sort_values(by='city')  # Sort rows by 'city' column
            week_group = week_group.drop(columns=['week', 'complaint_date'])

            for row in week_group.itertuples(index=False):
                ws.append(row)
                current_row += 1

        # Apply styles to cells
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Set column width
        for i, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            column = get_column_letter(i)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        max_col = ws.max_column
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=max_col, max_col=max_col):
            for cell in row:
                cell.value = None

        # Iterate through the rows and merge cells where the value in the first cell starts with a number
        for row in ws.iter_rows(min_row=4):
            value = row[0].value  # First cell value
            if value and str(value).startswith(('0', '1', '2', '3', '4', '5', '6', '7', '8', '9')):
                merge_range = f"A{row[0].row}:G{row[0].row}"
                ws.merge_cells(merge_range)

        # Set the last column fill to color FAD201
        max_col = ws.max_column
        column_letter = get_column_letter(max_col)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for cell in ws[column_letter]:
            cell.fill = fill

        # Add the value 'controle' in bold at row 3, column 7 with white background
        cell = ws.cell(row=3, column=7, value='controle')
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Save workbook to BytesIO object
        byte_data = BytesIO()
        wb.save(byte_data)

        return byte_data





