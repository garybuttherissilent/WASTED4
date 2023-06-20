from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border, Side
from io import BytesIO
from openpyxl.utils import get_column_letter


def create_styled_excel_street(df):
    # Create a workbook and add a worksheet to it
    wb = Workbook()
    ws = wb.active

    # Define styles
    header_font = Font(name='Calibri', bold=True, color=Color('000000'), size=16)  # Make text color black
    header_fill = PatternFill(start_color='FAD201', end_color='FAD201', fill_type='solid')
    subheader_font = Font(name='Calibri', bold=True, color=Color('000000'), size=12)  # Larger font for subheaders
    column_name_font = Font(name='Calibri', bold=True, color=Color('000000'))  # Bold font for column names
    cell_font = Font(name='Calibri', color=Color('000000'))
    cell_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # Set all borders

    # Create the main header
    street_address = df['street_address'].iloc[0]
    ws.merge_cells('A1:F2')
    ws['A1'] = 'COMPLAINTS FOR {}'.format(street_address)
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = thin_border


    df['controle'] = ''

    # Reorder the columns
    df = df[['route', 'complaint_date', 'address_number', 'modality', 'recipient', 'controle']]

    # Write column names to 3rd row
    for i, column_name in enumerate(df.columns, start=1):
        ws.cell(row=3, column=i, value=column_name)
        ws.cell(row=3, column=i).font = column_name_font
        ws.cell(row=3, column=i).border = thin_border

    # Write data to worksheet
    current_row = 4
    # Group by route
    for _, route_group in df.groupby('route'):
        route_value = route_group['route'].iloc[0]
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=1, value=route_value)
        cell.font = subheader_font
        cell.fill = header_fill
        cell.alignment = cell_alignment
        cell.border = thin_border
        current_row += 1

        route_group = route_group.sort_values(by='complaint_date')  # Sort rows by 'complaint_date' column

        for row in route_group.itertuples(index=False):
            ws.append(row)
            current_row += 1

    # Apply styles to cells
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Set column width
    for i, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(i)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save workbook to BytesIO object
    byte_data = BytesIO()
    wb.save(byte_data)

    return byte_data
