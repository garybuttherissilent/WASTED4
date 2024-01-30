import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Color, Border, Side
import datetime


def dagplanningmaker(file):
    # INLEZEN VAN SAP DATA
    # KOLOMMEN VERWIJDEREN, TOEVOEGEN
    # ORDENEN VOLGENS ROUTENUMMER EN KOPPELEN VOLGENS NAAM

    df = pd.read_excel(file)

    cols_to_drop = ['Orderdatum', 'Infotekst', 'Activiteit', 'Afvalfractie', 'Voertuig 2', 'Redencode',
                    'Omschr. redencode', 'Werknemer 4']

    # Get a list of the columns in the dataframe
    cols = df.columns.tolist()

    # Iterate through the list of columns to drop
    for col in cols_to_drop:
        # Check if the column exists in the dataframe
        if col in cols:
            # Drop the column if it exists
            df = df.drop(col, axis=1)

    df.dropna(thresh=3, inplace=True)
    df.rename(columns={'Werknemer 1': 'Chauffeur', 'Voertuig 1': 'Voertuig', 'Werknemer 2': 'Milieuwerker 1',
                       'Werknemer 3': 'Milieuwerker 2'}, inplace=True)
    df['new'] = df.groupby('Chauffeur')['Route'].transform('max')
    df = df.sort_values(by=['new', 'Chauffeur', 'Route'], ascending=[True, True, False]).drop('new', axis=1)
    df['Opmerkingen'] = np.where(df['Route'].str.contains("HHPMD"), "START PMD: ", "")
    df['OK/NOK'] = ""
    df['Gedaan'] = ""
    df['Binnen'] = ""
    df['OPS-code'] = ""

    # GEWENSTE VOLGORDE VAN DE FRACTIES

    huisvuil = df[df['Route'].str.contains("HHRESZ|HHPMD|HHRESK|HHRES AFR")]
    diftar = df[df['Route'].str.contains("HHGFZC|HHGFZZ|HHRESC|HHGFT AFR")]

    # Separate the HHCLUS rows from the papier DataFrame
    papier_hhclus = df[df['Route'].str.contains("HHCLUS")]
    papier_without_hhclus = df[df['Route'].str.contains("HHPAP") & ~df['Route'].str.contains("HHCLUS")]

    # Append the HHCLUS rows to the end of the papier DataFrame
    papier = pd.concat([papier_without_hhclus, papier_hhclus], axis=0)

    glas = df[df['Route'].str.contains("HHGLS")]
    grof = df[df['Route'].str.contains("HHGV|HHGHSV")]
    bedrijf = df[df['Route'].str.contains("BA")]
    trucks = df[df['Route'].str.contains("AA")]
    ondergronds = df[df['Route'].str.contains("OC")]



    # NIEUWE RIJEN TOEVOEGEN VOOR HEADERS

    new_row = pd.DataFrame(
        {'Voertuig': 'RESTAFVAL/PMD', 'Route': '', 'Chauffeur': '', 'Milieuwerker 1': '', 'Milieuwerker 2': '',
         'Opmerkingen': '', 'OK/NOK': '', 'Gedaan': '', 'Binnen': '', 'OPS-code': ''}, index=[0])
    new_row2 = pd.DataFrame(
        {'Voertuig': 'DIFTAR', 'Route': '', 'Chauffeur': '', 'Milieuwerker 1': '', 'Milieuwerker 2': '',
         'Opmerkingen': '', 'OK/NOK': '', 'Gedaan': '', 'Binnen': '', 'OPS-code':''}, index=[0])
    new_row3 = pd.DataFrame(
        {'Voertuig': 'PAPIER', 'Route': '', 'Chauffeur': '', 'Milieuwerker 1': '', 'Milieuwerker 2': '',
         'Opmerkingen': '', 'OK/NOK': '', 'Gedaan': '', 'Binnen': '', 'OPS-code':''}, index=[0])
    new_row4 = pd.DataFrame(
        {'Voertuig': 'GLAS', 'Route': '', 'Chauffeur': '', 'Milieuwerker 1': '', 'Milieuwerker 2': '',
         'Opmerkingen': '', 'OK/NOK': '', 'Gedaan': '', 'Binnen': '', 'OPS-code':''}, index=[0])
    new_row5 = pd.DataFrame(
        {'Voertuig': 'GROF VUIL', 'Route': '', 'Chauffeur': '', 'Milieuwerker 1': '', 'Milieuwerker 2': '',
         'Opmerkingen': '', 'OK/NOK': '', 'Gedaan': '', 'Binnen': '', 'OPS-code': ''}, index=[0])
    new_row6 = pd.DataFrame(
        {'Voertuig': 'BEDRIJFSAFVAL', 'Route': '', 'Chauffeur': '', 'Milieuwerker 1': '', 'Milieuwerker 2': '',
         'Opmerkingen': '', 'OK/NOK': '', 'Gedaan': '', 'Binnen': '', 'OPS-code': ''}, index=[0])
    new_row7 = pd.DataFrame(
        {'Voertuig': 'CONTAINERTRUCKS', 'Route': '', 'Chauffeur': '', 'Milieuwerker 1': '', 'Milieuwerker 2': '',
         'Opmerkingen': '', 'OK/NOK': '', 'Gedaan': '', 'Binnen': '', 'OPS-code':''}, index=[0])
    new_row8 = pd.DataFrame(
        {'Voertuig': 'ONDERGRONDS', 'Route': '', 'Chauffeur': '', 'Milieuwerker 1': '', 'Milieuwerker 2': '',
         'Opmerkingen': '', 'OK/NOK': '', 'Gedaan': '', 'Binnen': '', 'OPS-code': ''}, index=[0])


    # CONCATENEREN VAN HEADERS EN FRACTIES

    df = pd.concat(
        [new_row, huisvuil, new_row2, diftar, new_row3, papier, new_row4, glas, new_row5, grof, new_row6, bedrijf,
         new_row7, trucks, new_row8, ondergronds], axis=0)

    data = df.values.tolist()

    # OPSLAAN VOOR OPMAAK
    wb = Workbook()
    sh1 = wb.active

    # Iterate over the DataFrame and add the rows and columns to the Excel sheet
    for row in data:
        sh1.append(row)

    # Get the column names from the dataframe
    column_names = df.columns.tolist()
    sh1.insert_rows(1)

    # Write the column names to the first row of the worksheet
    for i, name in enumerate(column_names):
        sh1.cell(row=1, column=i + 1).value = name

    # Set the font of the first row to Calibri with a size of 14
    for cell in sh1[1]:
        cell.font = Font(name="Calibri", size=14, bold=True)

    # OPMAAK GROOTTE

    for row in sh1.iter_rows():
        for cell in row:
            cell.font = Font(name="Calibri", size=14)

    # Insert four new rows at the top of the worksheet
    sh1.insert_rows(1, 4)

    # HEADER VOOR DATUM


    today = datetime.date.today()
    week = today.isocalendar()[1]
    fdate = today.strftime("%A %d/%m/%Y")
    date_string = today.strftime('%d-%m-%Y')

    sh1['A1'] = fdate
    sh1['A1'].font = Font(bold=True, size=26)
    sh1['A1'].fill = PatternFill(start_color="FFDB58", end_color="FFDB58", fill_type="solid")
    sh1['A6'].font = Font(bold=True, size=18)
    sh1['A6'].fill = PatternFill(start_color="FCF55F", end_color="FCF55F", fill_type="solid")

    # ALL BORDERS

    for row in sh1.iter_rows():
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000'),
                                 left=Side(border_style='thin', color='FF000000'))

    # SETTING COLUMMNHEADERS BOLD AND BIG

    col11 = sh1['A5']
    col11.font = Font(bold=True, size=16)
    col22 = sh1['B5']
    col22.font = Font(bold=True, size=16)
    col33 = sh1['C5']
    col33.font = Font(bold=True, size=16)
    col44 = sh1['D5']
    col44.font = Font(bold=True, size=16)
    col55 = sh1['E5']
    col55.font = Font(bold=True, size=16)
    col66 = sh1['F5']
    col66.font = Font(bold=True, size=16)
    col77 = sh1['G5']
    col77.font = Font(bold=True, size=16)
    col88 = sh1['H5']
    col88.font = Font(bold=True, size=16)
    col99 = sh1['I5']
    col99.font = Font(bold=True, size=16)
    col1010 = sh1['J5']
    col1010.font = Font(bold=True, size=16)

    # KLEURENOPMAAK
    for rows in sh1.iter_rows(min_row=6, max_row=28, min_col=8, max_col=10):
        for cell in rows:
            cell.fill = PatternFill(start_color='FCF55F', end_color='FCF55F', fill_type="solid")

    for row in sh1.iter_rows():
        for cell in row:
            if cell.value == "START PMD: ":
                cell.font = Font(bold=True, size=14)
                varpmd = cell.row
                sh1.cell(row=varpmd, column=8).fill = PatternFill(start_color='87CEEB', end_color='87CEEB',
                                                                  fill_type="solid")
                sh1.cell(row=varpmd, column=9).fill = PatternFill(start_color='87CEEB', end_color='87CEEB',
                                                                  fill_type="solid")
                sh1.cell(row=varpmd, column=10).fill = PatternFill(start_color='87CEEB', end_color='87CEEB',
                                                                  fill_type="solid")
            if cell.value == 'DIFTAR' and week % 2 != 0:
                cell.value = 'DIFTAR//GFT'
                vardif = cell.row
                cell.font = Font(bold=True, size=18)
                cell.fill = PatternFill(start_color="B4C424", end_color="B4C424", fill_type="solid")
            if cell.value == 'DIFTAR' and week % 2 == 0:
                cell.value = 'DIFTAR//REST'
                vardif = cell.row
                cell.font = Font(bold=True, size=18)
                cell.fill = PatternFill(start_color="C4B454", end_color="C4B454", fill_type="solid")
            if cell.value == 'PAPIER':
                varpap = cell.row
                cell.font = Font(bold=True, size=18)
                cell.fill = PatternFill(start_color="FBCEB1", end_color="FBCEB1", fill_type="solid")
            if cell.value == 'GLAS':
                vargls = cell.row
                cell.font = Font(bold=True, size=18)
                cell.fill = PatternFill(start_color="FFAA33", end_color="FFAA33", fill_type="solid")
            if cell.value == 'GROF VUIL':
                vargrfv = cell.row
                cell.font = Font(bold=True, size=18)
                cell.fill = PatternFill(start_color="EEDC82", end_color="EEDC82", fill_type="solid")
            if cell.value == 'BEDRIJFSAFVAL':
                varba = cell.row
                cell.font = Font(bold=True, size=18)
            if cell.value == 'CONTAINERTRUCKS':
                varct = cell.row
                cell.font = Font(bold=True, size=18)
            if cell.value == 'ONDERGRONDS':
                varoc = cell.row
                cell.font = Font(bold=True, size=18)
            else:
                pass

    for rows in sh1.iter_rows(min_row=vardif, max_row=varpap, min_col=8, max_col=10):
        for cell in rows:
            if week % 2 != 0:
                cell.fill = PatternFill(start_color="B4C424", end_color="B4C424", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color='C4B454', end_color='C4B454', fill_type="solid")

    for rows in sh1.iter_rows(min_row=varpap, max_row=vargls, min_col=8, max_col=10):
        for cell in rows:
            cell.fill = PatternFill(start_color='FBCEB1', end_color='FBCEB1', fill_type="solid")

    for rows in sh1.iter_rows(min_row=vargls, max_row=vargrfv, min_col=8, max_col=10):
        for cell in rows:
            cell.fill = PatternFill(start_color="FFAA33", end_color="FFAA33", fill_type="solid")

    for rows in sh1.iter_rows(min_row=vargrfv, max_row=varba, min_col=8, max_col=10):
        for cell in rows:
            cell.fill = PatternFill(start_color="EEDC82", end_color="EEDC82", fill_type="solid")

    ### SPECIALE GEBEURTENISSEN

    for row in sh1.iter_rows():
        for cell in row:
            if cell.value == "HHPMDK402" or cell.value == "HHPMDK101":
                sh1.cell(row=7, column=6).value = None
                sh1.cell(row=8, column=6).value = None
            if cell.value == "HHRESZ506":
                sh1.cell(row=18, column=6).value = "START REST: "
                sh1.cell(row=18, column=6).font = Font(bold=True, size=14)
                sh1.cell(row=20, column=6).value = "START REST: "
                sh1.cell(row=20, column=6).font = Font(bold=True, size=14)
                sh1.cell(row=19, column=6).value = None
                sh1.cell(row=21, column=6).value = None

    # FRACTIES ANDERE DIENSTEN GRIJS KLEUREN

    lastrow = sh1.max_row
    for rows in sh1.iter_rows(min_row=varba, max_row=lastrow, min_col=0, max_col=10):
        for cell in rows:
            cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type="solid")

    # MERGING HEADERS AND CENTERING

    sh1.merge_cells('A1:J4')
    sh1.merge_cells('A6:J6')
    sh1.merge_cells(start_row=vardif, end_row=vardif, start_column=1, end_column=10)
    sh1.merge_cells(start_row=varpap, end_row=varpap, start_column=1, end_column=10)
    sh1.merge_cells(start_row=vargls, end_row=vargls, start_column=1, end_column=10)
    sh1.merge_cells(start_row=vargrfv, end_row=vargrfv, start_column=1, end_column=10)
    sh1.merge_cells(start_row=varba, end_row=varba, start_column=1, end_column=10)
    sh1.merge_cells(start_row=varct, end_row=varct, start_column=1, end_column=10)
    sh1.merge_cells(start_row=varoc, end_row=varoc, start_column=1, end_column=10)

    for row in range(1, sh1.max_row + 1):
        for col in range(1, sh1.max_column + 1):
            cell = sh1.cell(row, col)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # SIZE OF COLUMNS

    sh1.column_dimensions['A'].width = 15
    sh1.column_dimensions['B'].width = 20
    sh1.column_dimensions['C'].width = 30
    sh1.column_dimensions['D'].width = 30
    sh1.column_dimensions['E'].width = 30
    sh1.column_dimensions['F'].width = 35
    sh1.column_dimensions['G'].width = 10
    sh1.column_dimensions['H'].width = 13
    sh1.column_dimensions['I'].width = 13
    sh1.column_dimensions['J'].width = 13


    return wb


